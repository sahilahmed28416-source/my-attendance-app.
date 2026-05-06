import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import io
from datetime import datetime

# --- CONFIGURATION & UI ---
st.set_page_config(page_title="SafeTech Precast - Workforce Pro", layout="wide")

# Custom CSS for Professional Look
st.markdown("""
    <style>
    .main { background-color: #f5f7f9; }
    .stButton>button { width: 100%; border-radius: 5px; height: 3em; background-color: #1E3A8A; color: white; }
    .stSidebar { background-color: #e5eaf5; }
    h1 { color: #1E3A8A; font-family: 'Arial'; border-bottom: 2px solid #1E3A8A; }
    .status-box { padding: 10px; border-radius: 5px; margin: 5px 0; border-left: 5px solid #1E3A8A; background: white; }
    </style>
    """, unsafe_allow_html=True)

# Logo Handling
try:
    # Bhai, file ka naam 'logo.jpg' ya jo bhi aapki image hai wo rakho
    st.image("3810b91a7e82de2cb2273cb3494b93cbc0589111", width=120) 
except:
    st.sidebar.warning("Logo image not found. Upload 'logo.jpg' in the directory.")

st.title("🏗️ SAFETECH PRECAST - WORKFORCE MANAGEMENT")

uploaded_file = st.file_uploader("📂 Upload SafeTech Master Attendance File (.xlsx)", type=["xlsx"])

if uploaded_file:
    file_bytes = uploaded_file.read()
    
    # --- 1. SIDEBAR: REGISTRATION & WORKER STATUS ---
    st.sidebar.header("👤 Employee Management")
    
    with st.sidebar.expander("➕ Register New Employee", expanded=False):
        with st.form("new_emp_form"):
            e_id = st.text_input("Employee ID (T00...)")
            e_fname = st.text_input("First Name")
            e_lname = st.text_input("Last Name")
            e_desig = st.text_input("Designation")
            e_dept = st.text_input("Department")
            # Base Status (Aapne poocha tha - Shift fixed rahegi yahan)
            e_shift = st.selectbox("Assigned Base Shift", ["Day (D)", "Night (N)"])
            e_join = st.date_input("Joining Date", datetime.now())
            
            submit_w = st.form_submit_button("Register in All Months")

    if submit_w and e_id:
        with st.spinner("Updating all 12 sheets..."):
            wb = load_workbook(io.BytesIO(file_bytes))
            full_name = f"{e_fname} {e_lname}".strip()
            shift_code = "D" if "Day" in e_shift else "N"
            
            for month_name in wb.sheetnames:
                if month_name != "At Record":
                    ws = wb[month_name]
                    # Data Row 15 se shuru hai, check kar rahe hain next empty row
                    target_row = 15
                    while ws.cell(row=target_row, column=2).value is not None:
                        target_row += 1
                    
                    # Writing Data
                    ws.cell(row=target_row, column=1).value = target_row - 14 # Srl No
                    ws.cell(row=target_row, column=2).value = e_id        # B: ID
                    ws.cell(row=target_row, column=3).value = full_name   # C: Name
                    ws.cell(row=target_row, column=4).value = e_desig     # D: Desig
                    ws.cell(row=target_row, column=5).value = e_dept      # E: Dept
                    ws.cell(row=target_row, column=6).value = e_join      # F: Joining
                    # Column AL (38) ko hum Base Status shift ke liye use kar sakte hain
                    ws.cell(row=target_row, column=38).value = shift_code 
            
            out = io.BytesIO()
            wb.save(out)
            file_bytes = out.getvalue()
            st.sidebar.success(f"Employee {e_id} added successfully!")

    # --- 2. MAIN INTERFACE: ATTENDANCE SYSTEM ---
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    month_list = [m for m in xls.sheet_names if m != "At Record"]
    
    col1, col2 = st.columns([1, 3])
    with col1:
        st.info("📅 Select Context")
        sel_month = st.selectbox("Select Month", month_list)
        
        # Display Summary directly from the sheet (Rows 1-12)
        st.markdown("<div class='status-box'><b>Summary View Active</b><br>Rows 1-12 are preserved.</div>", unsafe_allow_html=True)

    # Reading the data (Row 13 header)
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sel_month, skiprows=12)
    df_raw.columns = [str(c).strip() for c in df_raw.columns]
    
    # Filter rows that have an Emp ID
    df_active = df_raw[df_raw['Emp ID'].notna()].copy()

    with col2:
        st.subheader(f"Register View: {sel_month} 2026")
        st.dataframe(df_active, use_container_width=True, hide_index=True)

    # --- 3. BULK ATTENDANCE & SHIFT LOGIC ---
    st.divider()
    st.header(f"⚡ Monthly Attendance Operation - {sel_month}")
    
    tab1, tab2 = st.tabs(["📋 Bulk Entry", "🔍 Individual Search"])
    
    with tab1:
        st.write("Paste employee IDs and apply status to specific dates.")
        b_col1, b_col2, b_col3, b_col4 = st.columns([2, 1, 1, 1])
        
        with b_col1:
            bulk_ids = st.text_area("Employee IDs (Paste from Excel or WhatsApp)", height=150, placeholder="T001\nT002\nT003")
        with b_col2:
            target_date = st.number_input("Enter Date (1-31)", 1, 31, value=1)
            # Fetch day name for user confirmation
            try:
                # Column G index 6 is Day 1
                day_name = df_raw.columns[6 + target_date - 1]
                st.write(f"Selected Day: **{day_name}**")
            except: pass
        with b_col3:
            shift_mode = st.radio("Shift Selection", ["Day Shift", "Night Shift"])
            st.caption("Filters status options based on shift.")
        with b_col4:
            if shift_mode == "Day Shift":
                status_list = ["DP", "DA", "L", "T", "WO", "ST"]
            else:
                status_list = ["NP", "NA", "L", "T", "WO", "ST"]
            final_status = st.selectbox("Select Status", status_list)

        if st.button("🚀 Apply & Save Attendance"):
            if not bulk_ids:
                st.error("Please enter Employee IDs.")
            else:
                wb = load_workbook(io.BytesIO(file_bytes))
                ws = wb[sel_month]
                id_list = [i.strip().upper() for i in bulk_ids.replace(',', '\n').split('\n') if i.strip()]
                
                # Logic: Column G is Day 1 (7th column)
                col_to_update = 6 + target_date
                
                updated = 0
                for r in range(14, ws.max_row + 1):
                    val = str(ws.cell(row=r, column=2).value).strip().upper()
                    if val in id_list:
                        ws.cell(row=r, column=col_to_update).value = final_status
                        # Visual feedback color
                        ws.cell(row=r, column=col_to_update).fill = PatternFill(start_color="CCFFCC", fill_type="solid")
                        updated += 1
                
                out = io.BytesIO()
                wb.save(out)
                file_bytes = out.getvalue()
                st.success(f"Success! {updated} records updated for Day {target_date} in {sel_month}.")
                st.rerun()

    with tab2:
        # Search feature for quick checks
        search_query = st.text_input("Search Worker by ID or Name")
        if search_query:
            search_results = df_active[df_active.apply(lambda row: search_query.lower() in str(row).lower(), axis=1)]
            st.table(search_results[['Srl No', 'Emp ID', 'Emp Name', 'Designation', 'Department']])

    # --- 4. EXPORT ---
    st.divider()
    st.download_button(
        label="📥 DOWNLOAD UPDATED SAFETECH MASTER FILE",
        data=file_bytes,
        file_name=f"SafeTech_Master_Attendance_Final_{datetime.now().strftime('%d_%m')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.markdown("""
        <div style='text-align:center; padding: 50px; border: 2px dashed #1E3A8A; border-radius:15px;'>
            <h3>Bhai, Master Attendance File Upload Kariye</h3>
            <p>Code automatically Row 13 headers aur har month ki sheets ko detect kar lega.</p>
        </div>
        """, unsafe_allow_html=True)
